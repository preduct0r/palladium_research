
import os
from pathlib import Path
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.styles.borders import Border, Side
import pandas as pd

# --- 1. Пути к файлам ------------------------------------------------
OUTPUT_PATH     = Path("Анализ статей (ИИ)_заполненный.xlsx")
DATA_ROOT       = Path("short_data")                     # корень со статьями
ANSWERS_DIRNAME = "answers"                        # подпапка с txt
SERVICE_EXCEL   = Path("Анализ статей (ИИ).xlsx")  # файл с правилами окраски

# --- 2. Функция для чтения цветовых правил из Excel ----------------
def load_color_rules(excel_file):
    """Загружает правила окраски из вкладки Service Excel файла"""
    try:
        df = pd.read_excel(excel_file, sheet_name='Service', header=None)
        color_rules = {}
        
        # Проверяем все столбцы начиная с D (индекс 3)
        for col_idx in range(3, df.shape[1]):
            col_data = df.iloc[:, col_idx].dropna()
            if len(col_data) > 1:  # если есть данные кроме заголовка
                header = col_data.iloc[0]
                options = col_data.iloc[1:].tolist()
                if options:  # если есть варианты
                    color_rules[header] = options
        
        return color_rules
    except Exception as e:
        print(f"Ошибка при загрузке правил окраски: {e}")
        return {}

# --- 3. Функция для определения цвета ячейки -----------------------
def get_cell_color(value, column_header, color_rules):
    """Определяет цвет ячейки на основе значения и правил окраски"""
    if column_header not in color_rules:
        return None
    
    options = color_rules[column_header]
    
    # Приводим значение к строке, убираем лишние пробелы и переводим в нижний регистр
    value_str = str(value).strip().lower()
    
    # Ищем точное совпадение (с учетом возможных различий в пробелах и регистре)
    for i, option in enumerate(options):
        option_str = str(option).strip().lower()
        if value_str == option_str:
            if i == 0:
                return "C6EFCE"  # Светло-зеленый - первый вариант
            elif i == len(options) - 1:
                return "F4CCCC"  # Светло-красный - последний вариант
            else:
                return "FFEB9C"  # Светло-желтый - промежуточные варианты
    
    return None  # Если значение не найдено в правилах

# --- 4. Загружаем правила окраски -----------------------------------
color_rules = load_color_rules(SERVICE_EXCEL)
print(f"Загружено правил окраски для {len(color_rules)} столбцов")

# --- 5. Сопоставление имён файлов → заголовки столбцов ---------------------------
map_names = {
    "article_name": "Название статьи",  # добавляем новый столбец для названия
    "doi": "Ссылка [DOI]",
    "date": "Дата публикации", 
    "journal": "Журнал",
    "tematic": "Направление (тематика)",
    "idea": "Идея",
    "technology": "Промышленная технология",
    "type": "Тип проекта",
    "palladium_novelty": "Новизна применения Pd",
    "technical_feasibility": "Научно-техническая реализуемость внедрения Pd",
    "technology_readiness_level": "Уровень готовности технологии с Pd",
    "technology_development_level": "Уровень развития технологии под которую делается продукт",
    "development_complexity": "Сложность разработки",
    "development_duration": "Длительность разработки", 
    "implementation_complexity": "Сложность внедрения",
    "commercial_potential": "Коммерческий потенциал внедрения Pd",
    "commercialization_potential": "Потенциал коммерциализации",
    "market_commercial_potential": "Уровень рыночного потенциала (коммерция)",
    "market_prospects": "Перспективность рынка (разработка)",
    "competitive_advantages": "Конкурентные преимущества Pd",
    "potential_consumption": "Потенциальное потребление Pd, кг",
    "decision": "Принятие решения",
    "comments": "Комментарии",
}

# --- 6. Создаём новую книгу --------------------------------------------
wb = Workbook()
ws = wb.active
ws.title = "Анализ статей"

# --- 7. Записываем строку заголовков -----------------------------------
headers = list(map_names.values())
for col, header in enumerate(headers, start=1):
    cell = ws.cell(row=1, column=col, value=header)
    # Форматирование заголовков
    cell.font = Font(bold=True, size=10)
    cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    cell.alignment = Alignment(wrapText=True, vertical="center", horizontal="center")

# --- 8. Настраиваем фиксированные размеры ---------------------------------------
COL_WIDTH   = 30     # ширина столбцов (увеличил для лучшей читаемости)
ROW_HEIGHT  = 120    # высота строк (увеличил для многострочного текста)

# Устанавливаем ширину столбцов
for col_idx in range(1, len(map_names) + 1):
    col_letter = get_column_letter(col_idx)
    ws.column_dimensions[col_letter].width = COL_WIDTH

# Высота заголовка
ws.row_dimensions[1].height = 60

# --- 9. Наполняем таблицу строками по статьям -----------------------------------
row_idx = 2   # начинаем с второй строки (после заголовков)

for article_dir in sorted(DATA_ROOT.iterdir()):
    if not article_dir.is_dir():
        continue
    
    answers_folder = article_dir / ANSWERS_DIRNAME
    if not answers_folder.is_dir():
        continue  # пропускаем, если подпапки 'answers' нет

    print(f"Обрабатываем: {article_dir.name}")
    
    for col_idx, (fname, column_header) in enumerate(map_names.items(), start=1):
        if fname == "article_name":
            # Для названия статьи используем имя папки, заменяя подчеркивания на пробелы
            value = article_dir.name.replace("_", " ")
        else:
            txt_path = answers_folder / f"{fname}.txt"
            
            if txt_path.exists():
                try:
                    value = txt_path.read_text(encoding="utf-8").strip()
                    # Ограничиваем длину текста для лучшего отображения
                    if len(value) > 1000:
                        value = value[:1000] + "..."
                except Exception as e:
                    print(f"Ошибка чтения файла {txt_path}: {e}")
                    value = ""
            else:
                value = ""
            
        cell = ws.cell(row=row_idx, column=col_idx, value=value)
        
        # Определяем цвет ячейки на основе правил
        cell_color = get_cell_color(value, column_header, color_rules)
        
        # Форматирование ячеек с данными
        cell.alignment = Alignment(
            wrapText=True, 
            vertical="top", 
            horizontal="left",
            shrinkToFit=True  # автоматически уменьшает шрифт если не помещается
        )
        cell.font = Font(size=9)
        
        # Применяем цветовую заливку если есть правило
        if cell_color:
            cell.fill = PatternFill(start_color=cell_color, end_color=cell_color, fill_type="solid")

    # Устанавливаем высоту строки
    ws.row_dimensions[row_idx].height = ROW_HEIGHT
    row_idx += 1

# --- 10. Добавляем границы для всех ячеек ------------------------------------
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

for row in ws.iter_rows(min_row=1, max_row=row_idx-1, min_col=1, max_col=len(map_names)):
    for cell in row:
        cell.border = thin_border

# --- 11. Сохраняем результат ------------------------------------------------------
wb.save(OUTPUT_PATH)
print(f"Готово! Файл сохранён как: {OUTPUT_PATH.resolve()}")
print(f"Обработано статей: {row_idx - 2}")
print(f"Столбцов: {len(map_names)}")
print(f"Применено цветовых правил для {len([h for h in headers if h in color_rules])} столбцов")



