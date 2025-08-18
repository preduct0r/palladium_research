
import os
from pathlib import Path
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.styles.borders import Border, Side

# --- 1. Пути к файлам ------------------------------------------------
OUTPUT_PATH     = Path("Анализ статей (ИИ)_long.xlsx")
DATA_ROOT       = Path("long_data")                     # корень со статьями
ANSWERS_DIRNAME = "answers"                        # подпапка с txt

# --- 2. Сопоставление имён файлов → заголовки столбцов ---------------------------
map_names = {
    "article_name": "Название статьи",  # добавляем новый столбец для названия
    "doi": "Ссылка [DOI]",
    "date": "Дата публикации", 
    "journal": "Журнал",
    "tematic": "Направление (тематика)",
    "idea": "Идея",
    "technology": "Промышленная технология",
    "type": "Тип проекта",
    "palladium_novelty": "Новизна применения Pd",
    "technical_feasibility": "Научно-техническая реализуемость внедрения Pd",
    "technology_readiness_level": "Уровень готовности технологии с Pd",
    "technology_development_level": "Уровень развития технологии под которую делается продукт",
    "development_complexity": "Сложность разработки",
    "development_duration": "Длительность разработки", 
    "implementation_complexity": "Сложность внедрения",
    "commercial_potential": "Коммерческий потенциал внедрения Pd",
    "commercialization_potential": "Потенциал коммерциализации",
    "market_commercial_potential": "Уровень рыночного потенциала (коммерция)",
    "market_prospects": "Перспективность рынка (разработка)",
    "competitive_advantages": "Конкурентные преимущества Pd",
    "potential_consumption": "Потенциальное потребление Pd, кг",
    "decision": "Принятие решения",
    "comments": "Комментарии",
}

# --- 3. Создаём новую книгу --------------------------------------------
wb = Workbook()
ws = wb.active
ws.title = "Анализ статей"

# --- 4. Записываем строку заголовков -----------------------------------
headers = list(map_names.values())
for col, header in enumerate(headers, start=1):
    cell = ws.cell(row=1, column=col, value=header)
    # Форматирование заголовков
    cell.font = Font(bold=True, size=10)
    cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    cell.alignment = Alignment(wrapText=True, vertical="center", horizontal="center")

# --- 5. Настраиваем фиксированные размеры ---------------------------------------
COL_WIDTH   = 30     # ширина столбцов (увеличил для лучшей читаемости)
ROW_HEIGHT  = 120    # высота строк (увеличил для многострочного текста)

# Устанавливаем ширину столбцов
for col_idx in range(1, len(map_names) + 1):
    col_letter = get_column_letter(col_idx)
    ws.column_dimensions[col_letter].width = COL_WIDTH

# Высота заголовка
ws.row_dimensions[1].height = 60

# --- 6. Наполняем таблицу строками по статьям -----------------------------------
row_idx = 2   # начинаем с второй строки (после заголовков)

for article_dir in sorted(DATA_ROOT.iterdir()):
    if not article_dir.is_dir():
        continue
    
    answers_folder = article_dir / ANSWERS_DIRNAME
    if not answers_folder.is_dir():
        continue  # пропускаем, если подпапки 'answers' нет

    print(f"Обрабатываем: {article_dir.name}")
    
    for col_idx, (fname, _) in enumerate(map_names.items(), start=1):
        if fname == "article_name":
            # Для названия статьи используем имя папки, заменяя подчеркивания на пробелы
            value = article_dir.name.replace("_", " ")
        else:
            txt_path = answers_folder / f"{fname}.txt"
            
            if txt_path.exists():
                try:
                    value = txt_path.read_text(encoding="utf-8").strip()
                    # Ограничиваем длину текста для лучшего отображения
                    if len(value) > 1000:
                        value = value[:1000] + "..."
                except Exception as e:
                    print(f"Ошибка чтения файла {txt_path}: {e}")
                    value = ""
            else:
                value = ""
            
        cell = ws.cell(row=row_idx, column=col_idx, value=value)
        # Форматирование ячеек с данными
        cell.alignment = Alignment(
            wrapText=True, 
            vertical="top", 
            horizontal="left",
            shrinkToFit=True  # автоматически уменьшает шрифт если не помещается
        )
        cell.font = Font(size=9)

    # Устанавливаем высоту строки
    ws.row_dimensions[row_idx].height = ROW_HEIGHT
    row_idx += 1

# --- 7. Добавляем границы для всех ячеек ------------------------------------
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

for row in ws.iter_rows(min_row=1, max_row=row_idx-1, min_col=1, max_col=len(map_names)):
    for cell in row:
        cell.border = thin_border

# --- 8. Сохраняем результат ------------------------------------------------------
wb.save(OUTPUT_PATH)
print(f"Готово! Файл сохранён как: {OUTPUT_PATH.resolve()}")
print(f"Обработано статей: {row_idx - 2}")
print(f"Столбцов: {len(map_names)}")



